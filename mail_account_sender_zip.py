import sys
import openpyxl
from pathlib import Path
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# 送信先の読み込み
wb_master = openpyxl.load_workbook("アカウント情報作成.xlsm")
ws_master = wb_master["送信先"]

account_list = []
# iter_rows:上から１行ずつデータを吟味しながら処理する
for row in ws_master.iter_rows(min_row=2):
    if row[0].value is None:
        break
    value_list = []
    for c in row:
        value_list.append(c.value)
    
    print(value_list)
    account_list.append(value_list)

# 請求書PDFのフォルダー
pdf_dir = "PDF"

# メーリングリスト
mailing_list = []

# フォルダーからのアカウント情報ZIPファイルを1つずつ取得する
for invoice in Path(pdf_dir).glob("*.zip"):
    # 「名前」 は、 ZIPファイルの拡張子を除いた部分
    account_id = invoice.stem
    # 該当するアカウントデータを「」から検索
    for account in account_list:
        if account_id == account[4]:
            # メーリングリストに「送信先データ」と
            # 「ZIPファイル」のパスを追加
            mailing_list.append([account, invoice])
            # 部署名、名前、メールアドレス、
            # ZIPファイルのパスを表示
            print(account[0], account[1], account[4], invoice)

# モード選択
print()  # 1行空ける
mode = input("モード選択（テスト＝test、本番=real）：")
# 本番以外はテスト
if mode != "real":
    test_mode = True
else:
    test_mode = False

# 送信確認
if test_mode:
    result = input("テストモードで自分宛てに送信します（続行＝yes、中止＝no）：")
else:
    result = input("本番モードで送信します（続行＝yes、中止＝no）：")

# 続行以外は中止
if result != "yes":
    print("プログラムを中止します")
    sys.exit()

# メール本文をファイルから読み込んでおく
text = open("mail_body.txt", encoding="utf-8")
body_temp = text.read()
text.close()

my_address = "iida@cbl.or.jp"

# SMTPサーバー
smtp_server = "smtp.cbl.or.jp"
port_number = 587

# ログイン情報
account = "iida"
password = "YUpZUB4GPzpP"

# SMTPサーバーに接続
server = smtplib.SMTP(smtp_server, port_number)
server.login(account, password)

# メーリングリストの顧客に1つずつメール送信
for data in mailing_list:
    account = data[0]
    pdf_file = data[1]

    # メッセージの準備
    msg = MIMEMultipart()
    # 件名、メールアドレスの設定
    msg["Subject"] = "【再送】メールパスワードの変更及びメールの取りこぼしの確認方法について"
    msg["From"] = my_address
    if test_mode:
        msg["To"] = my_address
    else:
        msg["To"] = account[3]

    # メール本文の追加
    body_text = body_temp.format(
        department=account[0],
        person=account[1],
        position=account[2]
    )
    body = MIMEText(body_text)
    msg.attach(body)
    # 添付ファイルの追加
    pdf = open(pdf_file, mode="rb")
    pdf_data = pdf.read()
    pdf.close()
    attach_file = MIMEApplication(pdf_data)
    attach_file.add_header("Content-Disposition", "attachment", filename=pdf_file.name)
    msg.attach(attach_file)

    # メール送信
    print("メール送信：", account[0], account[1], account[2])
    server.send_message(msg)

# SMTPサーバーとの接続を閉じる
server.quit()
